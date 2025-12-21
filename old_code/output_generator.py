"""
Module for writing the analysis results to an Excel file (CLI version).
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
logger = logging.getLogger(__name__)





def write_results_to_excel(monthly_params, analysis_results, output_path, raw_df=None):
	"""
	Writes five output blocks to a new Excel file: monthly parameters, seasonal reversion, volatility, trend, and key correlation coefficients.

	Args:
		monthly_params (dict): Monthly parameter results.
		analysis_results (dict): Analysis results with keys 'seasonal_reversion', 'seasonal_volatility', 'seasonal_trend', 'correlation_matrix'.
		output_path (str): Path to save the new Excel file.
	"""
	try:
		logger.debug(f"Writing results to Excel file: {output_path}")
		wb = Workbook()
		ws = wb.active
		ws.title = 'Parameters_Output'

		# Block 1: Long-Term Monthly Parameters
		ws['A1'] = 'Long-Term Monthly Parameters'
		monthly_df = pd.DataFrame.from_dict(monthly_params, orient='index')
		monthly_df.index.name = 'Month'
		monthly_df = monthly_df[['p_ww', 'p_wd', 'alpha', 'beta']]
		for r_idx, row in enumerate(dataframe_to_rows(monthly_df.reset_index(), index=False, header=True), start=2):
			for c_idx, value in enumerate(row, start=1):
				ws.cell(row=r_idx, column=c_idx, value=value)

		# Block 2: Reversion Rate (4x4)
		block2_start = r_idx + 3
		ws.cell(row=block2_start, column=1, value='Reversion Rate (by Season and Parameter)')
		reversion_df = analysis_results['seasonal_reversion']
		for r2_idx, row in enumerate(dataframe_to_rows(reversion_df.reset_index(), index=False, header=True), start=block2_start+1):
			for c_idx, value in enumerate(row, start=1):
				ws.cell(row=r2_idx, column=c_idx, value=value)

		# Block 3: Volatility (4x4)
		block3_start = r2_idx + 3
		ws.cell(row=block3_start, column=1, value='Volatility (by Season and Parameter)')
		volatility_df = analysis_results['seasonal_volatility']
		for r3_idx, row in enumerate(dataframe_to_rows(volatility_df.reset_index(), index=False, header=True), start=block3_start+1):
			for c_idx, value in enumerate(row, start=1):
				ws.cell(row=r3_idx, column=c_idx, value=value)

		# Block 4: Trend Slope (4x4)
		block4_start = r3_idx + 3
		ws.cell(row=block4_start, column=1, value='Trend Slope (by Season and Parameter)')
		trend_df = analysis_results['seasonal_trend']
		for r4_idx, row in enumerate(dataframe_to_rows(trend_df.reset_index(), index=False, header=True), start=block4_start+1):
			for c_idx, value in enumerate(row, start=1):
				ws.cell(row=r4_idx, column=c_idx, value=value)

		# Block 5: Correlation Coefficients (annual means)
		block5_start = r4_idx + 3
		ws.cell(row=block5_start, column=1, value='Correlation Coefficients (Annual Means)')
		corr = analysis_results['correlation_matrix']
		corr_pairs = [
			('PWW-PWD', corr.loc['PWW', 'PWD'] if 'PWW' in corr.index and 'PWD' in corr.columns else None),
			('PWW-ALPHA', corr.loc['PWW', 'ALPHA'] if 'PWW' in corr.index and 'ALPHA' in corr.columns else None),
			('PWD-ALPHA', corr.loc['PWD', 'ALPHA'] if 'PWD' in corr.index and 'ALPHA' in corr.columns else None),
			('ALPHA-BETA', 0.8),  # Hardcoded as specified
		]
		table5 = pd.DataFrame(corr_pairs, columns=['Parameter Pair', 'Correlation Coefficient'])
		for r5_idx, row in enumerate(dataframe_to_rows(table5, index=False, header=True), start=block5_start+1):
			for c_idx, value in enumerate(row, start=1):
				ws.cell(row=r5_idx, column=c_idx, value=value)


		# Block 6: Overall Data Coverage (to the right of main tables)
		stats_col = 8  # Place stats to the right of main tables
		ws.cell(row=1, column=stats_col, value='Data Coverage:')
		if raw_df is not None:
			error_codes = set([-99, -999, -9999, 999, 9999, 99999])
			df = raw_df.copy()
			df['precip'] = pd.to_numeric(df['precip'], errors='coerce')
			df.loc[df['precip'].isin(error_codes), 'precip'] = float('nan')
			total_days = len(df)
			valid_days = df['precip'].notna().sum()
			percent_complete = (valid_days / total_days) if total_days > 0 else 0
			ws.cell(row=1, column=stats_col+1, value=f"{round(percent_complete*100, 1)}%")
		else:
			ws.cell(row=1, column=stats_col+1, value='(Raw data not available)')

		wb.save(output_path)
		logger.debug(f"Results successfully written to {output_path}")
	except Exception as e:
		logger.exception(f"Error writing results to Excel: {e}")
		raise
